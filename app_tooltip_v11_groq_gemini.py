# -*- coding: utf-8 -*-
"""
Streamlit app - Predicción de producción de café (Wizard para producción)

✅ Fondo con imagen (robusto en múltiples contenedores Streamlit)
✅ Hero + subtítulo
✅ Indicador "Paso X de 3" con etiqueta del paso
✅ Wizard: 1) Plantilla  2) Subir archivo  3) Predicción + métricas opcionales
✅ Botones responden a 1 click (st.rerun() tras cambiar step)
✅ Sin sidebar; utilidades en "Modelo (opcional)" (expander visible tipo glass)
✅ Predicción blindada: usa SOLO features definidas por la plantilla (contrato)
✅ Métricas: MAE/RMSE con >=1 dato; R² con >=2 datos
✅ RMSE compatible (sin squared=False)
"""

import base64
import io
import inspect
import os
from math import sqrt
from pathlib import Path
from typing import Optional, List

import joblib
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.tree import export_text

# Import opcional para recomendaciones IA con Groq
try:
    from groq import Groq
except Exception:
    Groq = None

# Import opcional para recomendaciones IA con Gemini
try:
    from google import genai
except Exception:
    genai = None

# Import opcional: necesario si tu .joblib serializa una clase personalizada CLR
try:
    from clr_model import CLR  # noqa: F401
except Exception:
    pass


# -----------------------------
# Config general
# -----------------------------
st.set_page_config(
    page_title="Predicción de producción de café",
    layout="centered",
    initial_sidebar_state="collapsed",
)

BASE_DIR = Path(__file__).resolve().parent
MODEL_PATH = BASE_DIR / "best_model.joblib"

# Plantilla (contrato de columnas del modelo)
TEMPLATE_CANDIDATES = [
    BASE_DIR / "Plantilla-vacia-Coffee pests and diseases survey, Costa Rica, 2002-2003, Jacques Avelino.xlsx",
    BASE_DIR / "assets" / "plantilla_cultivo_cafe.xlsx",
    BASE_DIR / "assets" / "Plantilla-vacia-Coffee pests and diseases survey, Costa Rica, 2002-2003, Jacques Avelino.xlsx",
]

# Diccionario de variables (nombres amigables ES)
VAR_DICT_CANDIDATES = [
    BASE_DIR / "caracteristicas-modelo_con_descripciones_ES_final.xlsx",
    BASE_DIR / "assets" / "caracteristicas-modelo_con_descripciones_ES_final.xlsx",
]


OPTIONAL_METRIC_COL = "produccion_real"  # columna opcional para métricas (NO entra al modelo)


# -----------------------------
# Compatibilidad Streamlit: algunos widgets cambiaron use_container_width -> width
# -----------------------------
def _btn_kwargs():
    try:
        return {"width": "stretch"} if "width" in inspect.signature(st.button).parameters else {"use_container_width": True}
    except Exception:
        return {"use_container_width": True}


def _dl_kwargs():
    try:
        return {"width": "stretch"} if "width" in inspect.signature(st.download_button).parameters else {"use_container_width": True}
    except Exception:
        return {"use_container_width": True}


def _df_kwargs():
    # dataframe() aún usa use_container_width en muchas versiones; esto evita warnings
    return {"use_container_width": True}


BTN_KW = _btn_kwargs()
DL_KW = _dl_kwargs()
DF_KW = _df_kwargs()


def first_existing_path(candidates: List[Path]) -> Optional[Path]:
    for p in candidates:
        if p.exists():
            return p
    return None


# -----------------------------
# Imagen de fondo (detección robusta)
# -----------------------------
def find_bg_image(base_dir: Path) -> Optional[Path]:
    candidates: List[Path] = []
    candidates += list(base_dir.glob("cafe-web.*"))
    assets_dir = base_dir / "assets"
    if assets_dir.exists():
        candidates += list(assets_dir.glob("cafe-web.*"))

    priority = {".jpg": 0, ".jpeg": 1, ".png": 2, ".webp": 3}
    candidates = sorted(candidates, key=lambda p: priority.get(p.suffix.lower(), 99))
    return candidates[0] if candidates else None


BG_IMAGE_PATH = find_bg_image(BASE_DIR)
TEMPLATE_PATH = first_existing_path(TEMPLATE_CANDIDATES)
VAR_DICT_PATH = first_existing_path(VAR_DICT_CANDIDATES)


# -----------------------------
# UI helpers
# -----------------------------
def set_bg_corporate(image_path: Optional[Path]):
    if image_path is None or not image_path.exists():
        st.markdown(
            """
            <style>
            html, body, [data-testid="stAppViewContainer"], .stApp {
                background: linear-gradient(160deg, rgba(20,14,10,0.92), rgba(20,14,10,0.85)) !important;
                background-size: cover !important;
                background-position: center !important;
                background-attachment: fixed !important;
            }
            [data-testid="stHeader"] { background: rgba(0,0,0,0) !important; }
            [data-testid="stSidebar"] { display: none; }
            [data-testid="stSidebarCollapsedControl"] { display: none; }
            /* ===== MÉTRICAS VISIBLES ===== */
            div[data-testid="metric-container"] {
                background: rgba(255,255,255,0.10) !important;
                border: 1px solid rgba(255,255,255,0.15) !important;
                border-radius: 14px !important;
                padding: 12px 16px !important;
            }

            div[data-testid="metric-container"] label,
            div[data-testid="metric-container"] p {
                color: rgba(255,255,255,0.95) !important;
                font-weight: 700 !important;
            }

            div[data-testid="metric-container"] [data-testid="stMetricValue"],
            div[data-testid="metric-container"] [data-testid="stMetricValue"] * {
                color: #FFFFFF !important;
                font-weight: 900 !important;
                font-size: 2rem !important;
            }

            div[data-testid="metric-container"] [data-testid="stMetricDelta"],
            div[data-testid="metric-container"] [data-testid="stMetricDelta"] * {
                color: rgba(255,255,255,0.85) !important;
            }

            div[data-testid="metric-container"] div {
                color: #FFFFFF !important;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
        return

    data = base64.b64encode(image_path.read_bytes()).decode("utf-8")

    st.markdown(
        f"""
        <style>
        html, body {{
            height: 100%;
        }}

        [data-testid="stAppViewContainer"], .stApp {{
            background:
              linear-gradient(160deg, rgba(20,14,10,0.88), rgba(20,14,10,0.78)),
              url("data:image/jpg;base64,{data}") !important;
            background-size: cover !important;
            background-position: center !important;
            background-attachment: fixed !important;
        }}

        [data-testid="stHeader"] {{
            background: rgba(0,0,0,0) !important;
        }}

        h1, h2, h3 {{
            color: #FFFFFF !important;
            letter-spacing: -0.02em;
        }}
        p, span, label {{
            color: rgba(255,255,255,0.90) !important;
        }}

        [data-testid="stVerticalBlockBorderWrapper"] {{
            background: rgba(255,255,255,0.97);
            border-radius: 18px;
            padding: 22px;
            box-shadow: 0px 10px 28px rgba(0,0,0,0.28);
            border: 1px solid rgba(0,0,0,0.06);
        }}

        [data-testid="stVerticalBlockBorderWrapper"] h1,
        [data-testid="stVerticalBlockBorderWrapper"] h2,
        [data-testid="stVerticalBlockBorderWrapper"] h3,
        [data-testid="stVerticalBlockBorderWrapper"] p,
        [data-testid="stVerticalBlockBorderWrapper"] span,
        [data-testid="stVerticalBlockBorderWrapper"] label {{
            color: rgba(12,12,12,0.92) !important;
        }}

        div.stButton > button {{
            background: #6F4E37 !important;
            color: #FFFFFF !important;
            border: 1px solid rgba(255,255,255,0.15) !important;
            border-radius: 12px !important;
            padding: 0.65rem 1rem !important;
            font-weight: 700 !important;
        }}
        div.stButton > button:hover {{
            filter: brightness(1.05);
            transform: translateY(-1px);
        }}
        .stDownloadButton > button {{
            background: #6F4E37 !important;
            color: #FFFFFF !important;
            border-radius: 12px !important;
            font-weight: 700 !important;
        }}
        /* Refuerzo: download buttons (evita que queden blancos por wrappers) */
        div[data-testid="stDownloadButton"] > button,
        div[data-testid="stDownloadButton"] button,
        .stDownloadButton > button,
        .stDownloadButton button {{
            background: #6F4E37 !important;
            color: #FFFFFF !important;
            border: 1px solid rgba(255,255,255,0.15) !important;
            border-radius: 12px !important;
            font-weight: 700 !important;
        }}
        div[data-testid="stDownloadButton"] > button:hover,
        div[data-testid="stDownloadButton"] button:hover {{
            filter: brightness(1.05);
            transform: translateY(-1px);
        }}

        [data-testid="stFileUploader"] {{
            background: rgba(255,255,255,0.97);
            border-radius: 16px;
            padding: 12px 14px;
            border: 1px solid rgba(0,0,0,0.08);
            box-shadow: 0px 8px 22px rgba(0,0,0,0.14);
        }}
        [data-testid="stFileUploader"] * {{
            color: rgba(12,12,12,0.92) !important;
        }}
        [data-testid="stFileUploaderDropzone"] {{
            background: rgba(255,255,255,1) !important;
            border: 1px dashed rgba(0,0,0,0.20) !important;
        }}

        /* ✅ Expander visible como panel gris “glass” */
        [data-testid="stExpander"] {{
          background: rgba(255,255,255,0.14) !important;
          border: 1px solid rgba(255,255,255,0.18) !important;
          border-radius: 14px !important;
          backdrop-filter: blur(6px);
          margin-bottom: 10px;
        }}
        [data-testid="stExpander"] details > summary {{
          background: rgba(255,255,255,0.18) !important;
          border-radius: 14px !important;
          padding: 10px 12px !important;
        }}
        [data-testid="stExpander"] details > summary * {{
          color: rgba(255,255,255,0.92) !important;
          font-weight: 700 !important;
        }}
        [data-testid="stExpander"] details > div {{
          padding: 10px 12px 14px 12px !important;
        }}

        [data-testid="stSidebar"] {{ display: none; }}
        [data-testid="stSidebarCollapsedControl"] {{ display: none; }}

        /* ===== MÉTRICAS VISIBLES ===== */
        div[data-testid="metric-container"] {{
            background: rgba(255,255,255,0.10) !important;
            border: 1px solid rgba(255,255,255,0.15) !important;
            border-radius: 14px !important;
            padding: 12px 16px !important;
        }}

        div[data-testid="metric-container"] label,
        div[data-testid="metric-container"] p {{
            color: rgba(255,255,255,0.95) !important;
            font-weight: 700 !important;
        }}

        div[data-testid="metric-container"] [data-testid="stMetricValue"],
        div[data-testid="metric-container"] [data-testid="stMetricValue"] * {{
            color: #FFFFFF !important;
            font-weight: 900 !important;
            font-size: 2rem !important;
        }}

        div[data-testid="metric-container"] [data-testid="stMetricDelta"],
        div[data-testid="metric-container"] [data-testid="stMetricDelta"] * {{
            color: rgba(255,255,255,0.85) !important;
        }}

        div[data-testid="metric-container"] div {{
            color: #FFFFFF !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )



def inject_model_tooltips():
    """
    Tooltip real (hover sobre TODO el botón) para los 2 botones dentro del expander 'Modelo (opcional)'.

    Nota: NO usamos help=... porque en algunos temas/CSS altera el DOM y puede dejar botones sin estilo.
    Este hack agrega el atributo HTML title directamente sobre los botones renderizados.
    """
    js = """
    <script>
    const T1 = "Descarga el modelo entrenado (.joblib) para usarlo en Python o integrarlo en otra app.";
    const T2 = "Descarga un Excel con el resumen del modelo (reglas/estructura y parámetros) para documentación.";

    function apply() {
      try {
        // Busca botones por su texto (más robusto que asumir posiciones)
        const doc = window.parent.document;

        // Limita al expander cuyo summary contenga 'Modelo (opcional)'
        const expanders = doc.querySelectorAll('div[data-testid="stExpander"]');
        expanders.forEach(exp => {
          const summary = exp.querySelector('summary');
          const label = (summary && summary.innerText) ? summary.innerText.trim() : "";
          if (!label.includes("Modelo (opcional)")) return;

          const btns = exp.querySelectorAll('button');
          btns.forEach(b => {
            const txt = (b.innerText || "").trim();

            // OJO: los emojis pueden venir separados; usamos includes()
            if (txt.includes("Descargar mejor modelo") || txt.includes("Descargar mejor modelo (.joblib)") || txt.includes("Descargar mejor modelo (.joblib)")) {
              b.setAttribute("title", T1);
              b.style.cursor = "help";
            }
            if (txt.includes("Exportar info del modelo") || txt.includes("Exportar info del modelo (Excel)")) {
              b.setAttribute("title", T2);
              b.style.cursor = "help";
            }
          });
        });
      } catch (e) {
        // Silencioso: no queremos romper la app si el DOM cambia
      }
    }

    // Reintenta varias veces porque Streamlit re-renderiza componentes
    apply();
    const obs = new MutationObserver(apply);
    obs.observe(window.parent.document.body, { childList: true, subtree: true });
    setInterval(apply, 800);
    </script>
    """
    components.html(js, height=0, width=0)



def hero_panel():
    st.markdown(
        """
        <div style="
            padding:18px 20px;
            border-radius:18px;
            background: rgba(0,0,0,0.28);
            border: 1px solid rgba(255,255,255,0.12);
            backdrop-filter: blur(6px);
            margin-bottom: 14px;">
          <h1 style="margin:0; color:white;">Predicción de producción de café</h1>
          <p style="margin:6px 0 0 0; color:rgba(255,255,255,0.92); font-size:1.02rem;">
            Descarga la plantilla → Diligénciala → Súbela → Obtén predicciones (y métricas opcionales).
          </p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def progress_badge(step: int, total: int = 3):
    labels = {1: "Descargar plantilla", 2: "Subir archivo", 3: "Predicción"}
    st.markdown(
        f"""
        <div style="display:flex; gap:10px; align-items:center; margin: 6px 0 14px 0;">
          <div style="
              padding: 6px 10px;
              border-radius: 999px;
              background: rgba(255,255,255,0.16);
              border: 1px solid rgba(255,255,255,0.18);
              color: white;
              font-weight: 700;">
            Paso {step} de {total}
          </div>
          <div style="color: rgba(255,255,255,0.92); font-weight: 600;">
            {labels.get(step, "")}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


@st.cache_resource
def load_model(path: Path):
    return joblib.load(path)


@st.cache_data
def get_features_from_template(template_path: Path) -> List[str]:
    df_tpl = pd.read_excel(template_path)
    return [c for c in df_tpl.columns if c != OPTIONAL_METRIC_COL]

@st.cache_data
def load_variable_dictionary(dict_path: Optional[Path]):
    """Carga un diccionario de variables (ES) para nombres amigables.
    Espera columnas: nombre_original, nombre_amigable, descripcion_amigable
    Devuelve: (map_name, map_desc)
    """
    map_name = {}
    map_desc = {}
    if dict_path is None or (not dict_path.exists()):
        return map_name, map_desc

    try:
        df = pd.read_excel(dict_path)
    except Exception:
        return map_name, map_desc

    cols = {str(c).strip().lower(): c for c in df.columns}
    c_orig = cols.get("nombre_original")
    c_name = cols.get("nombre_amigable")
    c_desc = cols.get("descripcion_amigable")

    if c_orig is None:
        return map_name, map_desc

    for _, r in df.iterrows():
        k = str(r.get(c_orig, "")).strip()
        if not k:
            continue
        if c_name is not None:
            v = str(r.get(c_name, "")).strip()
            if v:
                map_name[k] = v
        if c_desc is not None:
            d = str(r.get(c_desc, "")).strip()
            if d:
                map_desc[k] = d

    return map_name, map_desc


def fmt_impact(v: float) -> str:
    """Formato compacto para impactos coef*valor (con signo)."""
    try:
        if v == 0:
            return "0"
        av = abs(v)
        if av >= 10000 or av < 0.01:
            return f"{v:.3e}"
        return f"{v:,.2f}"
    except Exception:
        return str(v)


def decision_path_rules_struct(_dt_model, x_row, feature_names: List[str]) -> List[str]:
    """Ruta del árbol como lista de reglas (una línea por regla), usando nombres ya mapeados."""
    if _dt_model is None:
        return []

    tree = _dt_model.tree_
    node_id = 0
    rules = []

    while tree.feature[node_id] != -2:  # hoja
        f_idx = int(tree.feature[node_id])
        f_name = feature_names[f_idx] if feature_names and f_idx < len(feature_names) else f"feature_{f_idx}"
        thr = float(tree.threshold[node_id])
        val = float(x_row[f_idx]) if f_idx < len(x_row) and (x_row[f_idx] == x_row[f_idx]) else float("nan")

        if val <= thr:
            rules.append(f"{f_name} ≤ {thr:.6g} (valor={val:.6g})")
            node_id = int(tree.children_left[node_id])
        else:
            rules.append(f"{f_name} > {thr:.6g} (valor={val:.6g})")
            node_id = int(tree.children_right[node_id])

    return rules


def topk_features_elasticnet_struct(_model, cluster_id: int, x_row, feature_names: List[str], map_name: dict, k: int = 5):
    """Top-k por |coef*valor|. Incluye impacto = coef*valor (con signo) y nombre amigable."""
    if not hasattr(_model, "best_reg_models") or _model.best_reg_models is None:
        return []

    try:
        reg = _model.best_reg_models[int(cluster_id)]
    except Exception:
        reg = None

    if reg is None or not hasattr(reg, "coef_"):
        return []

    coef = reg.coef_
    mlen = min(len(coef), len(x_row))

    vals = [0.0 if (x_row[i] != x_row[i]) else float(x_row[i]) for i in range(mlen)]
    impacts = [float(coef[i]) * vals[i] for i in range(mlen)]   # coef*valor con signo
    contrib = [abs(impacts[i]) for i in range(mlen)]            # |coef*valor|

    top_idx = sorted(range(mlen), key=lambda i: contrib[i], reverse=True)[:k]

    items = []
    for i in top_idx:
        orig = feature_names[i] if feature_names and i < len(feature_names) else f"feature_{i}"
        items.append({
            "feature_original": orig,
            "feature_name": map_name.get(orig, orig),
            "valor": vals[i],
            "coef": float(coef[i]),
            "impacto": float(impacts[i]),
            "contrib_abs": float(contrib[i]),
        })
    return items


def build_agro_paragraphs(parcela_n: int, cluster_id: Optional[int], path_rules: List[str], top_items: List[dict]) -> tuple[str, str]:
    """(descripcion, recomendacion) técnica y amigable por parcela.
    - Solo nombres amigables
    - Recomendación incluye impacto (coef*valor) por variable
    """
    c_txt = "NA" if cluster_id is None or cluster_id < 0 else str(int(cluster_id))

    rules_inline = "; ".join(path_rules[:3]) if path_rules else "sin reglas disponibles"
    top_inline = ", ".join([it["feature_name"] for it in top_items[:5]]) if top_items else "sin variables dominantes disponibles"

    descripcion = (
        f"Cultivo {parcela_n}: clasificado en la clase {c_txt}. "
        f"La clasificación estuvo determinada principalmente por: {rules_inline}. "
        f"Las variables con mayor influencia en la predicción fueron: {top_inline}."
    )

    rec_parts = []
    for it in top_items[:3]:
        rec_parts.append(f"{it['feature_name']} (impacto={fmt_impact(it['impacto'])})")

    if rec_parts:
        recomendacion = (
            "Recomendación técnica: prioriza el monitoreo y ajuste del manejo en las variables de mayor impacto: "
            + ", ".join(rec_parts)
            + ". "
            "Si el impacto es negativo, esa condición puede estar actuando como limitante; "
            "si es positivo, podría estar favoreciendo el rendimiento bajo el perfil del cluster."
        )
    else:
        recomendacion = (
            "Recomendación técnica: no fue posible identificar variables dominantes para este cultivo. "
            "Verifica valores numéricos y que el modelo tenga coeficientes por cluster."
        )

    return descripcion, recomendacion


def generar_recomendacion_ia_groq(cultivo_n, cluster_id, descripcion, recomendacion_base, explicacion_tecnica, model_id="llama-3.1-8b-instant"):
    """Genera una recomendación técnica con Groq usando la información ya calculada por la app."""
    if Groq is None:
        return "Groq no está instalado en este entorno. Ejecuta: pip install groq"

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        return "No se encontró la variable de entorno GROQ_API_KEY en este equipo."

    prompt = f"""
Actúa como un ingeniero agrónomo especialista en café.

Redacta una recomendación técnica breve, clara y profesional para el cultivo indicado.
Debes basarte únicamente en la información suministrada.
No inventes variables, no recomiendes acciones que no estén sustentadas por los datos y evita generalidades innecesarias.

Información del cultivo:
- Cultivo: {cultivo_n}
- Cluster asignado: {cluster_id}
- Descripción técnica: {descripcion}
- Recomendación técnica base: {recomendacion_base}
- Explicación técnica: {explicacion_tecnica}

Instrucciones:
1. Responde en español.
2. Entrega un solo párrafo de máximo 120 palabras.
3. Enfócate en variables de mayor impacto, posibles limitantes y acciones agronómicas concretas.
4. Si una variable aparece con impacto negativo, trátala como posible limitante.
5. Si una variable aparece con impacto positivo, trátala como factor favorable a sostener o potenciar.
6. No menciones que eres una IA.
"""

    try:
        client = Groq(api_key=api_key)
        response = client.chat.completions.create(
            model=model_id,
            messages=[
                {"role": "system", "content": "Eres un ingeniero agrónomo especialista en café."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            max_tokens=220,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"No fue posible generar la recomendación IA con Groq: {e}"




def generar_recomendacion_ia_gemini(cultivo_n, cluster_id, descripcion, recomendacion_base, explicacion_tecnica, model_id="gemini-3-flash-preview"):
    """Genera una recomendación técnica con Gemini usando la información ya calculada por la app."""
    if genai is None:
        return "No se encontró la librería google-genai en este entorno. Ejecuta: pip install google-genai"

    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        return "No se encontró la variable de entorno GEMINI_API_KEY (o GOOGLE_API_KEY) en este equipo."

    prompt = f"""
Actúa como un ingeniero agrónomo especialista en café.

Redacta una recomendación técnica breve, clara y profesional para el cultivo indicado.
Debes basarte únicamente en la información suministrada.
No inventes variables, no recomiendes acciones que no estén sustentadas por los datos y evita generalidades innecesarias.

Información del cultivo:
- Cultivo: {cultivo_n}
- Cluster asignado: {cluster_id}
- Descripción técnica: {descripcion}
- Recomendación técnica base: {recomendacion_base}
- Explicación técnica: {explicacion_tecnica}

Instrucciones:
1. Responde en español.
2. Entrega un solo párrafo de máximo 120 palabras.
3. Enfócate en variables de mayor impacto, posibles limitantes y acciones agronómicas concretas.
4. Si una variable aparece con impacto negativo, trátala como posible limitante.
5. Si una variable aparece con impacto positivo, trátala como factor favorable a sostener o potenciar.
6. No menciones que eres una IA.
"""

    try:
        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(
            model=model_id,
            contents=prompt,
        )
        return (response.text or "").strip() if hasattr(response, "text") else str(response)
    except Exception as e:
        return f"No fue posible generar la recomendación IA con Gemini: {e}"


def construir_exportables(out, clusters, model, X_np, features, var_dict_path, include_ia=False, ia_provider="groq", ia_model="llama-3.1-8b-instant"):
    """Construye el Excel base o el Excel con recomendaciones IA."""
    map_name, _map_desc = load_variable_dictionary(var_dict_path)
    feature_names_orig = list(features)
    feature_names_friendly = [map_name.get(f, f) for f in feature_names_orig]

    rows = []
    informe_rows = []
    ia_rows = []

    dt = getattr(model, "best_dt_model", None)
    n_rows = int(out.shape[0])

    for i in range(n_rows):
        # Acepta tanto 'cluster' como 'clase'
        if "cluster" in out.columns and pd.notna(out.iloc[i]["cluster"]):
            c_id = int(out.iloc[i]["cluster"])
        elif "clase" in out.columns and pd.notna(out.iloc[i]["clase"]):
            c_id = int(out.iloc[i]["clase"])
        else:
            c_id = -1

        # Hoja técnica
        path_txt = describe_decision_path(dt, X_np[i], feature_names_orig)
        top_txt = top5_features_elasticnet(model, c_id, X_np[i], feature_names_orig) if c_id >= 0 else "Clase no disponible."

        msg = (
            "Cultivo " + str(i + 1) + "\n"
            + "Clase asignada: " + (str(c_id) if c_id >= 0 else "NA") + "\n\n"
            + "Condiciones del árbol de decisión (ruta):\n" + str(path_txt) + "\n\n"
            + "Top 5 características con mayor impacto en la predicción (|coef·valor|):\n" + str(top_txt)
        )
        rows.append({"explicacion": msg})

        # Hoja amigable
        if c_id >= 0:
            path_rules = decision_path_rules_struct(dt, X_np[i], feature_names_friendly)
            top_items = topk_features_elasticnet_struct(model, c_id, X_np[i], feature_names_orig, map_name, k=5)
        else:
            path_rules = []
            top_items = []

        descripcion, recomendacion = build_agro_paragraphs(i + 1, c_id, path_rules, top_items)

        # Por si aún sale "Parcela" en algún caso
        descripcion = descripcion.replace("Parcela", "Cultivo")

        informe_rows.append({
            "descripcion": descripcion,
            "recomendacion_tecnica": recomendacion,
        })

        if include_ia:
            if ia_provider == "gemini":
                recomendacion_ia = generar_recomendacion_ia_gemini(
                    cultivo_n=i + 1,
                    cluster_id=(c_id if c_id >= 0 else "NA"),
                    descripcion=descripcion,
                    recomendacion_base=recomendacion,
                    explicacion_tecnica=msg,
                    model_id=ia_model,
                )
            else:
                recomendacion_ia = generar_recomendacion_ia_groq(
                    cultivo_n=i + 1,
                    cluster_id=(c_id if c_id >= 0 else "NA"),
                    descripcion=descripcion,
                    recomendacion_base=recomendacion,
                    explicacion_tecnica=msg,
                    model_id=ia_model,
                )

            ia_rows.append({
                "cultivo": i + 1,
                "recomendacion_ia": recomendacion_ia,
            })

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="predicciones")

        from openpyxl.styles import Alignment, Font, PatternFill

        ws_pred = writer.book["predicciones"]

        # Encabezados
        for cell in ws_pred[1]:
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.font = Font(bold=True)

        # Primeras dos columnas en gris
        fill_grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_letter in ["A", "B"]:
            for cell in ws_pred[col_letter]:
                cell.fill = fill_grey
                cell.alignment = align

        # Hoja explicaciones por cultivo
        pd.DataFrame(rows).to_excel(writer, sheet_name="explicaciones_por_cultivo", index=False)
        ws = writer.book["explicaciones_por_cultivo"]
        for cell in ws["A"]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 110

        # Hoja informe por cultivo
        pd.DataFrame(informe_rows).to_excel(writer, sheet_name="informe_por_cultivo", index=False)
        ws2 = writer.book["informe_por_cultivo"]
        for cell in ws2["A"]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in ws2["B"]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.column_dimensions["A"].width = 80
        ws2.column_dimensions["B"].width = 95

        # Hoja recomendaciones IA
        if include_ia:
            ia_sheet = "recomendaciones_gemini_por_cultivo" if ia_provider == "gemini" else "recomendaciones_groq_por_cultivo"
            pd.DataFrame(ia_rows).to_excel(writer, sheet_name=ia_sheet, index=False)
            ws3 = writer.book[ia_sheet]

            for cell in ws3["A"]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for cell in ws3["B"]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            ws3.column_dimensions["A"].width = 10
            ws3.column_dimensions["B"].width = 110

            for cell in ws3[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    return buffer.getvalue()


@st.cache_data
def build_model_info_excel(_model, features: Optional[List[str]]) -> bytes:
    """Exporta un Excel con:
    - info_general: variables y descripción desde caracteristicas-modelo_con_descripciones_ES_final.xlsx
    - arbol_decision: reglas del árbol con 'class' -> 'clase'
    - clase_XX_coeficientes: una hoja por cada clase
    """

    # -------- info_general desde el archivo de variables
    if VAR_DICT_PATH and VAR_DICT_PATH.exists():
        try:
            df_general_raw = pd.read_excel(VAR_DICT_PATH)

            # Detectar columnas disponibles sin romper otros procesos
            cols = {str(c).strip().lower(): c for c in df_general_raw.columns}

            col_nombre = (
                cols.get("nombre amigable")
                or cols.get("nombre_amigable")
                or cols.get("nombre")
                or cols.get("nombre_original")
                or cols.get("nombre_variable")
            )

            col_desc = (
                cols.get("descripción amigable")
                or cols.get("descripcion amigable")
                or cols.get("descripcion_amigable")
                or cols.get("descripción")
                or cols.get("descripcion")
                or cols.get("descripcion_variable")
            )

            if col_nombre and col_desc:
                df_general = df_general_raw[[col_nombre, col_desc]].copy()
                df_general.columns = ["nombre_variable", "descripcion_variable"]
            else:
                df_general = pd.DataFrame(
                    {
                        "nombre_variable": ["ERROR"],
                        "descripcion_variable": [
                            "No se encontraron columnas válidas en caracteristicas-modelo_con_descripciones_ES_final.xlsx"
                        ],
                    }
                )

        except Exception as e:
            df_general = pd.DataFrame(
                {
                    "nombre_variable": ["ERROR"],
                    "descripcion_variable": [
                        f"No fue posible leer caracteristicas-modelo_con_descripciones_ES_final.xlsx: {e}"
                    ],
                }
            )
    else:
        df_general = pd.DataFrame(
            {
                "nombre_variable": ["ERROR"],
                "descripcion_variable": [
                    "No se encontró caracteristicas-modelo_con_descripciones_ES_final.xlsx"
                ],
            }
        )

    # -------- arbol_decision
    if hasattr(_model, "best_dt_model"):
        try:
            if (
                features
                and hasattr(_model.best_dt_model, "n_features_in_")
                and len(features) == int(_model.best_dt_model.n_features_in_)
            ):
                tree_rules = export_text(_model.best_dt_model, feature_names=list(features))
            elif features:
                tree_rules = export_text(_model.best_dt_model, feature_names=list(features))
            else:
                tree_rules = export_text(_model.best_dt_model)
        except Exception:
            tree_rules = export_text(_model.best_dt_model)

        # Reemplazar class por clase
        tree_rules = tree_rules.replace("class:", "clase:")
        tree_rules = tree_rules.replace("class ", "clase ")
        tree_rules = tree_rules.replace("class=", "clase=")

        df_tree = pd.DataFrame({"reglas": tree_rules.split("\n")})
    else:
        df_tree = pd.DataFrame({"reglas": ["El modelo no tiene atributo best_dt_model."]})

    # -------- coeficientes por clase
    rows = []
    if hasattr(_model, "best_reg_models"):
        for i, regr in enumerate(_model.best_reg_models):
            if regr is None or not hasattr(regr, "coef_"):
                continue

            for j, coef in enumerate(regr.coef_):
                rows.append(
                    {
                        "clase": i,
                        "indice_caracteristica": j,
                        "nombre_caracteristica": (
                            features[j] if features and j < len(features) else f"X{j}"
                        ),
                        "coeficiente": float(coef),
                        "activo": bool(coef != 0),
                        "intercepto": float(getattr(regr, "intercept_", float("nan"))),
                    }
                )

    df_coefs = (
        pd.DataFrame(rows)
        if rows
        else pd.DataFrame(
            {
                "nota": [
                    "El modelo no expone best_reg_models o no tiene coeficientes disponibles."
                ]
            }
        )
    )

    # -------- escribir Excel
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_general.to_excel(writer, sheet_name="info_general", index=False)
        df_tree.to_excel(writer, sheet_name="arbol_decision", index=False)

        if "clase" in df_coefs.columns:
            try:
                clases = sorted([int(c) for c in df_coefs["clase"].dropna().unique().tolist()])
            except Exception:
                clases = df_coefs["clase"].dropna().unique().tolist()

            for c in clases:
                df_c = df_coefs[df_coefs["clase"] == c].copy()
                sheet = f"clase_{int(c):02d}_coeficientes"
                df_c.to_excel(writer, sheet_name=sheet[:31], index=False)

    return buffer.getvalue()


def validate_features(df: pd.DataFrame, features: List[str]) -> List[str]:
    return [c for c in features if c not in df.columns]


def safe_numeric_series(s: pd.Series) -> pd.Series:
    ss = s.astype(str).str.strip()
    has_comma = ss.str.contains(",", regex=False)
    has_dot = ss.str.contains(r"\.", regex=True)

    ss = ss.where(~(has_comma & has_dot), ss.str.replace(".", "", regex=False))
    ss = ss.where(~has_comma, ss.str.replace(",", ".", regex=False))

    ss = ss.str.replace("$", "", regex=False).str.replace("COP", "", regex=False).str.replace(" ", "", regex=False)
    return pd.to_numeric(ss, errors="coerce")


def get_cluster_assignments(_model, X_np):
    """Devuelve el cluster asignado por el árbol de decisión si existe."""
    if hasattr(_model, "best_dt_model") and _model.best_dt_model is not None:
        try:
            return _model.best_dt_model.predict(X_np)
        except Exception:
            return None
    return None


def describe_decision_path(_dt_model, x_row, feature_names: List[str]) -> str:
    """Construye una explicación legible de la ruta (reglas) que siguió el árbol para una fila."""
    if _dt_model is None:
        return "No hay árbol de decisión disponible en el modelo."

    tree = _dt_model.tree_
    node_id = 0
    rules = []

    while tree.feature[node_id] != -2:  # -2 => hoja
        f_idx = int(tree.feature[node_id])
        f_name = feature_names[f_idx] if feature_names and f_idx < len(feature_names) else f"feature_{f_idx}"
        thr = float(tree.threshold[node_id])

        val = float(x_row[f_idx]) if f_idx < len(x_row) and (x_row[f_idx] == x_row[f_idx]) else float("nan")

        if val <= thr:
            rules.append(f"- {f_name} ≤ {thr:.6g}  (valor={val:.6g})")
            node_id = int(tree.children_left[node_id])
        else:
            rules.append(f"- {f_name} > {thr:.6g}  (valor={val:.6g})")
            node_id = int(tree.children_right[node_id])

    return "\n".join(rules) if rules else "No fue posible reconstruir la ruta del árbol."


def top5_features_elasticnet(_model, cluster_id: int, x_row, feature_names: List[str]) -> str:
    """Devuelve las 5 variables con mayor impacto local en la predicción, usando |coef*valor|."""
    if not hasattr(_model, "best_reg_models") or _model.best_reg_models is None:
        return "No hay modelos de regresión por clase disponibles."

    try:
        reg = _model.best_reg_models[int(cluster_id)]
    except Exception:
        reg = None

    if reg is None or not hasattr(reg, "coef_"):
        return "No hay ElasticNet disponible para esta clase."

    coef = reg.coef_
    m = min(len(coef), len(x_row))

    vals = [0.0 if (x_row[i] != x_row[i]) else float(x_row[i]) for i in range(m)]
    impacto = [float(coef[i]) * vals[i] for i in range(m)]
    contrib = [abs(impacto[i]) for i in range(m)]

    top_idx = sorted(range(m), key=lambda i: contrib[i], reverse=True)[:5]

    out_lines = []
    for i in top_idx:
        name = feature_names[i] if feature_names and i < len(feature_names) else f"feature_{i}"
        out_lines.append(
            f"- {name}: valor={vals[i]:.6g}, coef={float(coef[i]):.6g}, "
            f"coef·valor={impacto[i]:.6g}, |coef·valor|={contrib[i]:.6g}"
        )

    return "\n".join(out_lines) if out_lines else "No fue posible calcular las 5 características con mayor impacto."


# -----------------------------
# Start app
# -----------------------------
set_bg_corporate(BG_IMAGE_PATH)
hero_panel()
inject_model_tooltips()

if "step" not in st.session_state:
    st.session_state.step = 1


# ✅ 1 click: cambiamos step y forzamos rerun
def go_next():
    st.session_state.step = min(3, st.session_state.step + 1)
    st.rerun()


def go_prev():
    st.session_state.step = max(1, st.session_state.step - 1)
    st.rerun()


if not MODEL_PATH.exists():
    st.error(
        f"No encuentro el modelo en: {MODEL_PATH}\n\n"
        "Asegúrate de tener `best_model.joblib` en la misma carpeta que `app.py`."
    )
    st.stop()

model = load_model(MODEL_PATH)

FEATURES: Optional[List[str]] = None
if TEMPLATE_PATH and TEMPLATE_PATH.exists():
    FEATURES = get_features_from_template(TEMPLATE_PATH)

with st.expander("Modelo (opcional)"):
    c1, c2 = st.columns(2)

    with c1:
        st.download_button(
            "⬇️ Descargar mejor modelo (.joblib)",
            data=MODEL_PATH.read_bytes(),
            file_name="best_model.joblib",
            mime="application/octet-stream",
            **DL_KW,
        )

    with c2:
        info_bytes = build_model_info_excel(model, FEATURES)
        st.download_button(
            "📄 Exportar info del modelo (Excel)",
            data=info_bytes,
            file_name="info_modelo_CLR.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            **DL_KW,
        )


# -----------------------------
# STEP 1
# -----------------------------
if st.session_state.step == 1:
    progress_badge(1)

    st.subheader("Paso 1 — Descarga la plantilla")
    st.write("Descarga el Excel, completa los datos de tu cultivo y luego súbelo en el Paso 2.")

    if TEMPLATE_PATH and TEMPLATE_PATH.exists():
        st.download_button(
            "📥 Descargar plantilla (.xlsx)",
            data=TEMPLATE_PATH.read_bytes(),
            file_name="plantilla_cultivo_cafe.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            **DL_KW,
        )
        st.caption(f"Plantilla base: {TEMPLATE_PATH.name}")
        st.info(
            f"En la plantilla se puede agregar más de una fila que representará información de uno o más cultivos según corresponda."
        )
        st.info(
            f"Si quieres métricas, puedes agregar una columna opcional llamada "
            f"`{OPTIONAL_METRIC_COL}` con el valor real de pruducción. Esa columna NO se usa en la predicción."
        )
    else:
        st.warning(
            "No encontré la plantilla en el proyecto.\n\n"
            "Colócala junto a este archivo o en la carpeta `assets/`."
        )
        st.info("Aun así puedes continuar, pero la app no podrá validar columnas por plantilla.")

    col1, col2 = st.columns([1, 1])
    with col2:
        if st.button("Siguiente ➡️", **BTN_KW):
            go_next()


# -----------------------------
# STEP 2
# -----------------------------
elif st.session_state.step == 2:
    progress_badge(2)

    st.subheader("Paso 2 — Sube la plantilla diligenciada")
    uploaded = st.file_uploader("Cargar Excel (.xlsx)", type=["xlsx"])

    if uploaded is None:
        st.info("Sube tu archivo para continuar.")
        c1, _ = st.columns([1, 1])
        with c1:
            if st.button("⬅️ Atrás", **BTN_KW):
                go_prev()
    else:
        df = pd.read_excel(uploaded)
        # Si el usuario sube otro archivo, reiniciamos el editor para evitar estados viejos
        if st.session_state.get("uploaded_name") != getattr(uploaded, "name", None):
            st.session_state["uploaded_name"] = getattr(uploaded, "name", None)
            # Limpia estados relacionados (si existen)
            for k in ["df_pred"]:
                if k in st.session_state:
                    del st.session_state[k]

        st.success("✅ Archivo cargado correctamente.")

        total_filas = int(df.shape[0])
        total_cols = int(df.shape[1])
        st.markdown(f"**Filas cargadas:** {total_filas} &nbsp;&nbsp;|&nbsp;&nbsp; **Columnas:** {total_cols}")
        st.write("Vista previa / edición (los cambios se usarán en la predicción):")

        editar_todo = st.checkbox("Editar todas las filas (puede ser lento si son muchas)", value=False)
        if editar_todo:
            df_to_edit = df.copy()
        else:
            max_n = min(200, total_filas) if total_filas > 0 else 0
            n_show = st.slider("Filas a mostrar/editar", min_value=1, max_value=max_n, value=min(50, max_n), step=1) if max_n else 0
            df_to_edit = df.head(n_show).copy() if n_show else df.head(0).copy()
            st.caption("Tip: si necesitas editar una fila que no aparece, activa *Editar todas las filas*.")

        edited = st.data_editor(
            df_to_edit,
            num_rows="fixed",
            key="editor_df",
            **DF_KW,
        )

        # Aplicamos las ediciones al dataframe completo
        if editar_todo:
            df_final = edited.copy()
        else:
            df_final = df.copy()
            if len(edited) > 0:
                df_final.iloc[: len(edited), :] = edited.values

        # ✅ Este es el DF que se usará en el Paso 3
        st.session_state.df_input = df_final

        if FEATURES:
            missing = validate_features(df, FEATURES)
            extra = [c for c in df.columns if c not in FEATURES and c != OPTIONAL_METRIC_COL]
            if missing:
                st.error(f"Faltan columnas requeridas por el modelo: {missing}")
            else:
                if extra:
                    st.warning(f"Columnas extra detectadas (se ignorarán para predicción): {extra}")
        else:
            st.warning(
                "No se pudo cargar la plantilla para validar columnas. "
                "La predicción intentará usar todas las columnas excepto "
                f"`{OPTIONAL_METRIC_COL}`."
            )

        c1, c2 = st.columns([1, 1])
        with c1:
            if st.button("⬅️ Atrás", **BTN_KW):
                go_prev()

        can_next = True
        if FEATURES:
            can_next = len(validate_features(st.session_state.df_input, FEATURES)) == 0

        with c2:
            if st.button("Siguiente ➡️", disabled=not can_next, **BTN_KW):
                go_next()

        if FEATURES and not can_next:
            st.info("Completa las columnas faltantes para habilitar 'Siguiente'.")


# -----------------------------
# STEP 3
# -----------------------------
else:
    progress_badge(3)

    st.subheader("Paso 3 — Predicción")

    if "df_input" not in st.session_state:
        st.warning("No hay archivo cargado. Vuelve al Paso 2.")
        if st.button("⬅️ Atrás", **BTN_KW):
            go_prev()
        st.stop()

    df = st.session_state.df_input.copy()

    features = FEATURES if FEATURES else [c for c in df.columns if c != OPTIONAL_METRIC_COL]
    missing = validate_features(df, features)
    if missing:
        st.error(f"Faltan columnas requeridas para predecir: {missing}")
        if st.button("⬅️ Atrás", **BTN_KW):
            go_prev()
        st.stop()

    st.write("✅ Columnas listas para predicción.")
    with st.expander(f"Ver columnas usadas en el modelo ({len(features)})"):
        st.dataframe(pd.DataFrame({"columna": features}), height=240, **DF_KW)


    if st.button("⚙️ Generar predicciones", **BTN_KW):
       X = df[features].copy()
       X_np = X.to_numpy()
       y_pred = model.predict(X_np)

       # DataFrame interno
       out = df.copy()
       clusters = get_cluster_assignments(model, X_np)
       if clusters is not None:
           out["cluster"] = clusters
       out["prediccion"] = y_pred

       # Mantener interno para otros procesos
       st.session_state.df_pred = out

       st.success("Predicciones generadas.")
       st.write("Vista previa:")

       # Exportable / visual
       out_export = out.copy()
       if "cluster" in out_export.columns:
          out_export = out_export.rename(columns={"cluster": "clase"})

       # Reordenar columnas: clase y prediccion primero
       front = [c for c in ["clase", "prediccion"] if c in out_export.columns]
       rest = [c for c in out_export.columns if c not in front]
       out_export = out_export[front + rest]

       st.dataframe(out_export.head(20), **DF_KW)

       # Excel base
       excel_base_bytes = construir_exportables(
           out=out_export,
           clusters=clusters,
           model=model,
           X_np=X_np,
           features=features,
           var_dict_path=VAR_DICT_PATH,
           include_ia=False,
           ia_provider="groq",
       )
       st.session_state["excel_base_bytes"] = excel_base_bytes

       st.download_button(
           "⬇️ Descargar Excel con predicciones",
           data=st.session_state.get("excel_base_bytes", b""),
           file_name="predicciones.xlsx",
           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
           **DL_KW,
       )

    if "df_pred" in st.session_state:
        st.markdown("---")
        st.subheader("Recomendaciones con IA (Groq)")
        st.caption("Esta opción usa la variable de entorno GROQ_API_KEY y puede tardar más, porque genera una recomendación por cultivo usando un modelo en línea.")
        # Modelo fijo
        ia_model = "llama-3.3-70b-versatile"

        c_ia1, c_ia2 = st.columns([1, 1])
        with c_ia1:
            st.text_input(
                "Modelo Groq",
                value=ia_model,
                disabled=True
            )
                        
        with c_ia2:
            max_ia = int(st.session_state.df_pred.shape[0]) if hasattr(st.session_state.df_pred, "shape") else 1
            n_ia = st.number_input(
                "Número de cultivos a procesar con IA",
                min_value=1,
                max_value=max_ia,
                value=min(10, max_ia),
                step=1,
                key="groq_n_rows",
            )

        if st.button("🧠 Generar recomendaciones con Groq IA", **BTN_KW):
            if Groq is None:
                st.error("No se encontró la librería groq. Instala con: pip install groq")
            elif not os.environ.get("GROQ_API_KEY"):
                st.error("No se encontró la variable de entorno GROQ_API_KEY en este equipo.")
            else:
                out_ia = st.session_state.df_pred.copy().head(int(n_ia))
                X_ia = st.session_state.df_input[features].copy().head(int(n_ia))
                X_ia_np = X_ia.to_numpy()
                clusters_ia = out_ia["cluster"].to_numpy() if "cluster" in out_ia.columns else None

                # Versión exportable: cluster -> clase
                out_ia_export = out_ia.copy()
                if "cluster" in out_ia_export.columns:
                    out_ia_export = out_ia_export.rename(columns={"cluster": "clase"})

                front = [c for c in ["clase", "prediccion"] if c in out_ia_export.columns]
                rest = [c for c in out_ia_export.columns if c not in front]
                out_ia_export = out_ia_export[front + rest]

                with st.spinner("Generando recomendaciones con IA..."):
                    excel_ia_bytes = construir_exportables(
                        out=out_ia_export,
                        clusters=clusters_ia,
                        model=model,
                        X_np=X_ia_np,
                        features=features,
                        var_dict_path=VAR_DICT_PATH,
                        include_ia=True,
                        ia_provider="groq",
                        ia_model=ia_model,
                    )
                    st.session_state["excel_ia_bytes"] = excel_ia_bytes
                st.success("Recomendaciones IA generadas correctamente.")

        if "excel_ia_bytes" in st.session_state:
            st.download_button(
                "⬇️ Descargar Excel con recomendaciones IA",
                data=st.session_state["excel_ia_bytes"],
                file_name="predicciones_con_recomendaciones_Groq_ia.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                **DL_KW,
            )


        st.markdown("---")
        st.subheader("Recomendaciones con IA (Gemini)")
        st.caption("Esta opción usa la variable de entorno GEMINI_API_KEY (o GOOGLE_API_KEY) y puede tardar más, porque genera una recomendación por cultivo usando un modelo en línea.")
        
        # Modelo fijo
        gemini_model = "gemini-3-flash-preview"

        g_ia1, g_ia2 = st.columns([1, 1])
        with g_ia1:
            st.text_input(
                "Modelo Gemini",
                value=gemini_model,
                disabled=True
            )

        with g_ia2:
            max_gem = int(st.session_state.df_pred.shape[0]) if hasattr(st.session_state.df_pred, "shape") else 1
            n_gem = st.number_input(
                "Número de cultivos a procesar con Gemini",
                min_value=1,
                max_value=max_gem,
                value=min(10, max_gem),
                step=1,
                key="gemini_n_rows",
            )

        if st.button("🧠 Generar recomendaciones con Gemini IA", **BTN_KW):
            if genai is None:
                st.error("No se encontró la librería google-genai. Instala con: pip install google-genai")
            elif not (os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")):
                st.error("No se encontró la variable de entorno GEMINI_API_KEY (o GOOGLE_API_KEY) en este equipo.")
            else:
                out_gem = st.session_state.df_pred.copy().head(int(n_gem))
                X_gem = st.session_state.df_input[features].copy().head(int(n_gem))
                X_gem_np = X_gem.to_numpy()
                clusters_gem = out_gem["cluster"].to_numpy() if "cluster" in out_gem.columns else None

                # Versión exportable: cluster -> clase
                out_gem_export = out_gem.copy()
                if "cluster" in out_gem_export.columns:
                    out_gem_export = out_gem_export.rename(columns={"cluster": "clase"})                

                front = [c for c in ["clase", "prediccion"] if c in out_gem_export.columns]
                rest = [c for c in out_gem_export.columns if c not in front]
                out_gem_export = out_gem_export[front + rest]

                with st.spinner("Generando recomendaciones con Gemini..."):
                    excel_gem_bytes = construir_exportables(
                        out=out_gem_export,
                        clusters=clusters_gem,
                        model=model,
                        X_np=X_gem_np,
                        features=features,
                        var_dict_path=VAR_DICT_PATH,
                        include_ia=True,
                        ia_provider="gemini",
                        ia_model=gemini_model,
                    )
                    st.session_state["excel_gemini_bytes"] = excel_gem_bytes
                st.success("Recomendaciones Gemini generadas correctamente.")

        if "excel_gemini_bytes" in st.session_state:
            st.download_button(
                "⬇️ Descargar Excel con recomendaciones Gemini",
                data=st.session_state["excel_gemini_bytes"],
                file_name="predicciones_con_recomendaciones_Gemini_ia.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                **DL_KW,
            )

    st.markdown("---")
    st.subheader("Métricas (opcional)")

    if OPTIONAL_METRIC_COL not in df.columns:
        st.info(
            f"Si quieres calcular métricas, agrega una columna opcional llamada `{OPTIONAL_METRIC_COL}` "
            "en tu Excel con el valor real. Esa columna NO se usa en la predicción."
        )
    else:
        if "df_pred" not in st.session_state:
            st.warning("Primero genera predicciones para calcular métricas.")
        else:
            out = st.session_state.df_pred
            y_true = safe_numeric_series(out[OPTIONAL_METRIC_COL])
            y_pred = safe_numeric_series(out["prediccion"])

            mask = (~y_true.isna()) & (~y_pred.isna())
            n = int(mask.sum())

            if n < 1:
                st.warning(
                    f"No hay valores numéricos válidos en `{OPTIONAL_METRIC_COL}` para calcular métricas. "
                    "Revisa el formato (coma/punto, separador de miles)."
                )
            else:
                if st.button("📊 Calcular métricas", **BTN_KW):
                    yt = y_true[mask].to_numpy()
                    yp = y_pred[mask].to_numpy()

                    mae = mean_absolute_error(yt, yp)
                    rmse = sqrt(mean_squared_error(yt, yp))

                    c1m, c2m = st.columns(2)
                    c1m.metric("MAE", f"{mae:,.4f}")
                    c2m.metric("RMSE", f"{rmse:,.4f}")

                    if n >= 2:
                        r2 = r2_score(yt, yp)
                        st.metric("R²", f"{r2:,.4f}")
                    else:
                        st.info("R² requiere al menos 2 registros con valores reales.")

    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("⬅️ Atrás", **BTN_KW):
            go_prev()
    with c2:
        if st.button("Reiniciar ✅", **BTN_KW):
            for k in ["df_input", "df_pred"]:
                if k in st.session_state:
                    del st.session_state[k]
            st.session_state.step = 1
            st.rerun()